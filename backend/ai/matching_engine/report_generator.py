import io
import datetime
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_match_report_xlsx(job: Dict[str, Any], candidates: List[Dict[str, Any]], anonymize: bool = False) -> bytes:
    """
    Generates a professionally formatted Excel workbook summarizing candidate match results.
    
    :param job: Dict containing job details (title, skills_required, etc.)
    :param candidates: List of dicts representing candidates and their score metrics
    :param anonymize: If True, candidate names are replaced with "Candidate {id/index}"
    :return: Bytes of the Excel file
    """
    wb = Workbook()
    
    # Define styles
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1B365D")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="5A6A85")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True, color="1C2D42")
    normal_font = Font(name=font_family, size=11, color="1C2D42")
    italic_font = Font(name=font_family, size=10, italic=True, color="5A6A85")
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    accent_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # Score band fills & fonts
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    green_font = Font(name=font_family, size=11, bold=True, color="155724")
    
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    yellow_font = Font(name=font_family, size=11, bold=True, color="856404")
    
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    red_font = Font(name=font_family, size=11, bold=True, color="721C24")
    
    # Skills checklist fills & fonts
    check_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    check_font = Font(name=font_family, size=11, bold=True, color="137333")
    cross_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
    cross_font = Font(name=font_family, size=11, bold=True, color="C5221F")
    
    thin_border_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom_border = Border(
        left=thin_border_side, right=thin_border_side, 
        top=thin_border_side, bottom=Side(border_style="double", color="1B365D")
    )
    
    # ----------------------------------------------------
    # Sheet 1: Summary Dashboard
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Overview"
    ws_summary.sheet_view.showGridLines = True
    
    # Title
    ws_summary["A1"] = f"Recruiter Batch Match Report"
    ws_summary["A1"].font = title_font
    ws_summary["A2"] = f"Generated on {datetime.date.today().strftime('%B %d, %Y')} | Powered by CareerForge AI"
    ws_summary["A2"].font = subtitle_font
    
    # Job Info Card
    ws_summary["A4"] = "Job Profile Information"
    ws_summary["A4"].font = Font(name=font_family, size=13, bold=True, color="1B365D")
    
    job_labels = [
        ("Job Title:", job.get("title", "Unknown Role")),
        ("Required Skills:", ", ".join(job.get("skills_required", []))),
        ("Experience Level:", job.get("experience_level", "Not Specified")),
        ("Total Applicants Analyzed:", len(candidates))
    ]
    
    for idx, (label, val) in enumerate(job_labels, start=5):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"A{idx}"].font = bold_font
        ws_summary[f"B{idx}"] = val
        ws_summary[f"B{idx}"].font = normal_font
        ws_summary[f"B{idx}"].alignment = Alignment(wrap_text=True)
        ws_summary[f"A{idx}"].border = thin_border
        ws_summary[f"B{idx}"].border = thin_border
    
    # Quick Statistics Card
    ws_summary["A11"] = "Hiring Pipeline Insights"
    ws_summary["A11"].font = Font(name=font_family, size=13, bold=True, color="1B365D")
    
    # Calculate stats
    scores = [c.get("overall_score", c.get("score", 0)) for c in candidates]
    avg_score = sum(scores) / len(scores) if scores else 0
    top_score = max(scores) if scores else 0
    strong_fits = sum(1 for s in scores if s >= 0.8)
    potential_fits = sum(1 for s in scores if 0.6 <= s < 0.8)
    
    stat_labels = [
        ("Top Match Score:", f"{top_score * 100:.1f}%" if top_score <= 1.0 else f"{top_score:.1f}%"),
        ("Average Match Score:", f"{avg_score * 100:.1f}%" if avg_score <= 1.0 else f"{avg_score:.1f}%"),
        ("Strong Fits (Score ≥ 80%):", strong_fits),
        ("Potential Fits (Score 60%-79%):", potential_fits)
    ]
    
    for idx, (label, val) in enumerate(stat_labels, start=12):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"A{idx}"].font = bold_font
        ws_summary[f"B{idx}"] = val
        ws_summary[f"B{idx}"].font = normal_font
        ws_summary[f"A{idx}"].border = thin_border
        ws_summary[f"B{idx}"].border = thin_border
    
    # Auto-adjust Summary columns
    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 65
    
    # ----------------------------------------------------
    # Sheet 2: Candidate Rankings
    # ----------------------------------------------------
    ws_rankings = wb.create_sheet(title="Rankings")
    ws_rankings.sheet_view.showGridLines = True
    ws_rankings.freeze_panes = "A2"
    
    headers = [
        "Rank", "Candidate Name", "Overall Score", 
        "Skill Match %", "Experience Match %", "Cultural Fit %",
        "Missing Skills", "Recommendation"
    ]
    
    # Write headers
    for col_idx, text in enumerate(headers, start=1):
        cell = ws_rankings.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Sort candidates by overall score descending
    def get_score(cand):
        return cand.get("overall_score", cand.get("score", 0))
        
    sorted_candidates = sorted(candidates, key=get_score, reverse=True)
    
    for row_idx, cand in enumerate(sorted_candidates, start=2):
        rank = row_idx - 1
        
        # Anonymization logic
        raw_name = cand.get("name", cand.get("candidate_name", f"Candidate {cand.get('candidate_id', rank)}"))
        if anonymize:
            cand_name = f"Candidate {rank:02d} (Anonymized)"
        else:
            cand_name = raw_name
            
        score = get_score(cand)
        skill_alignment = cand.get("skill_alignment", cand.get("skill_match", 0))
        experience_fit = cand.get("experience_fit", cand.get("experience_match", 0))
        cultural_fit = cand.get("cultural_fit", 0.8) # default fallback
        
        # Missing skills heuristics
        required_skills = set(job.get("skills_required", []))
        candidate_skills = set(cand.get("skills", []))
        missing_skills = list(required_skills - candidate_skills)
        missing_skills_str = ", ".join(missing_skills) if missing_skills else "None"
        
        # Recommendations
        recs = cand.get("recommendations", [])
        if not recs and isinstance(cand.get("reasons"), list):
            recs = cand.get("reasons")
        recs_str = "; ".join(recs) if recs else "Good candidate, evaluate in interviews."
        
        # Apply score normalization (e.g. convert 0.85 -> 85%)
        pct_score = score * 100 if score <= 1.0 else score
        pct_skill = skill_alignment * 100 if skill_alignment <= 1.0 else skill_alignment
        pct_exp = experience_fit * 100 if experience_fit <= 1.0 else experience_fit
        pct_culture = cultural_fit * 100 if cultural_fit <= 1.0 else cultural_fit
        
        row_data = [
            rank,
            cand_name,
            pct_score / 100.0,
            pct_skill / 100.0,
            pct_exp / 100.0,
            pct_culture / 100.0,
            missing_skills_str,
            recs_str
        ]
        
        is_alt = (row_idx % 2 == 1)
        row_fill = alt_row_fill if is_alt else PatternFill(fill_type=None)
        
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_rankings.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.fill = row_fill
            
            # Format types
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="left")
            elif col_idx in [3, 4, 5, 6]:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
                
                # Apply conditional color styles to Overall Score (col 3)
                if col_idx == 3:
                    if val >= 0.8:
                        cell.fill = green_fill
                        cell.font = green_font
                    elif val >= 0.6:
                        cell.fill = yellow_fill
                        cell.font = yellow_font
                    else:
                        cell.fill = red_fill
                        cell.font = red_font
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # Auto-adjust column widths for Rankings
    for col in ws_rankings.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # handle formats & strings
            val_str = str(cell.value or '')
            if cell.number_format == "0.0%" and isinstance(cell.value, float):
                val_str = f"{cell.value * 100:.1f}%"
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Add buffer but cap the width for missing skills/recs to prevent giant columns
        col_width = min(max(max_len + 3, 10), 45)
        ws_rankings.column_dimensions[col_letter].width = col_width
    ws_rankings.row_dimensions[1].height = 28

    # ----------------------------------------------------
    # Sheet 3: Skills Matrix
    # ----------------------------------------------------
    ws_matrix = wb.create_sheet(title="Skills Matrix")
    ws_matrix.sheet_view.showGridLines = True
    ws_matrix.freeze_panes = "B2"
    
    req_skills = job.get("skills_required", [])
    if not req_skills:
        # Fallback to general skills gathered from candidates
        all_skills = set()
        for cand in candidates:
            all_skills.update(cand.get("skills", []))
        req_skills = list(all_skills)[:10]  # Cap at 10 to keep it readable
        
    matrix_headers = ["Candidate Name"] + req_skills
    
    # Write Matrix Headers
    for col_idx, skill in enumerate(matrix_headers, start=1):
        cell = ws_matrix.cell(row=1, column=col_idx, value=skill)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    for row_idx, cand in enumerate(sorted_candidates, start=2):
        rank = row_idx - 1
        raw_name = cand.get("name", cand.get("candidate_name", f"Candidate {cand.get('candidate_id', rank)}"))
        if anonymize:
            cand_name = f"Candidate {rank:02d} (Anonymized)"
        else:
            cand_name = raw_name
            
        # Write Candidate Name in Column A
        cell_name = ws_matrix.cell(row=row_idx, column=1, value=cand_name)
        cell_name.font = bold_font
        cell_name.border = thin_border
        cell_name.fill = alt_row_fill if row_idx % 2 == 1 else PatternFill(fill_type=None)
        
        cand_skills_lower = {s.lower() for s in cand.get("skills", [])}
        
        for col_idx, skill in enumerate(req_skills, start=2):
            cell = ws_matrix.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            
            if skill.lower() in cand_skills_lower:
                cell.value = "✓"
                cell.font = check_font
                cell.fill = check_fill
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = "✗"
                cell.font = cross_font
                cell.fill = cross_fill
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths for Skills Matrix
    for col in ws_matrix.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_width = min(max(max_len + 4, 12), 30)
        ws_matrix.column_dimensions[col_letter].width = col_width
    ws_matrix.row_dimensions[1].height = 28

    # Save to byte stream
    file_stream = io.BytesIO()
    wb.save(file_stream)
    return file_stream.getvalue()
